from flask import Flask, request, jsonify, render_template
from openpyxl import Workbook, load_workbook
from sentence_transformers import SentenceTransformer
from sklearn.metrics.pairwise import cosine_similarity
import numpy as np
import os
import datetime
from waitress import serve  # Importing Waitress
import webbrowser

app = Flask(__name__)

EXCEL_FILE = "tickets.xlsx"
SIMILARITY_THRESHOLD = 0.2  # You can tune this

# Load SentenceTransformer model once
model = SentenceTransformer('all-MiniLM-L6-v2')
print("✅ SentenceTransformer model loaded.")


def initialize_excel():
    """Create Excel file with headers if not exists."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Ticket ID", "Module", "Resolver", "Title", "Issue", "Solution", "Closed Date"])
        wb.save(EXCEL_FILE)
        print("✅ Excel file initialized.")


def load_data():
    """Load all ticket data from Excel file."""
    tickets = []
    if not os.path.exists(EXCEL_FILE):
        return tickets

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        tickets.append({
            "ticketID": row[0],
            "module": row[1],
            "resolver": row[2],
            "title": row[3],
            "issue": row[4],
            "solution": row[5],
            "closingDate": row[6]
        })
    return tickets


@app.route('/')
def index():
    current_date = datetime.date.today()
    tickets = load_data()
    return render_template('mainpg.html', current_date=current_date, tickets=tickets)


@app.route("/submit", methods=["POST"])
def submit_ticket():
    data = request.json
    if not data:
        return jsonify({"error": "No data provided."}), 400

    initialize_excel()

    new_ticket = [
        data.get("ticketID", ""),
        data.get("module", ""),
        data.get("resolver", ""),
        data.get("title", ""),
        data.get("issue", ""),
        data.get("solution", ""),
        data.get("closingDate", "")
    ]

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append(new_ticket)
    wb.save(EXCEL_FILE)

    return jsonify({"message": "Ticket saved successfully."}), 200


@app.route("/get_tickets", methods=["GET"])
def get_tickets():
    tickets = load_data()
    return jsonify(tickets)


@app.route("/search", methods=["GET"])
def search_ticket():
    query = request.args.get("q", "").strip()
    print(f"🔍 Received query: {query}")  # Debug

    tickets = load_data()

    if not query:
        print("ℹ️ Empty query, returning all tickets.")
        return jsonify(tickets)

    # Combine all ticket fields into a searchable text string
    ticket_texts = [
        f"{t['ticketID']} {t['module']} {t['resolver']} {t['title']} {t['issue']} {t['solution']} {t['closingDate']}"
        for t in tickets
    ]

    # Encode query and ticket texts
    query_embedding = model.encode([query])
    ticket_embeddings = model.encode(ticket_texts)

    # Compute cosine similarity scores
    similarities = cosine_similarity(query_embedding, ticket_embeddings)[0]
    print("📊 Similarity scores:", similarities)

    # Filter results above threshold
    results = [
        (score, ticket) for score, ticket in zip(similarities, tickets)
        if score >= SIMILARITY_THRESHOLD
    ]
    results.sort(key=lambda x: x[0], reverse=True)
    print(f"✅ Found {len(results)} matching tickets.")

    return jsonify([ticket for score, ticket in results])


if __name__ == "__main__":
    initialize_excel()
    url = "http://127.0.0.1:8080"
    print(f"✅ App running at: {url}")
    webbrowser.open_new(url)
    serve(app, host="0.0.0.0", port=8080)